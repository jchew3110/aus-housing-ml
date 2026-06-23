"""
Parsers for ABS Excel and RBA CSV time-series files.

ABS Excel format quirks handled here:
- Row 0: series descriptions (column names, semicolon-padded)
- Rows 1-9: metadata (skip for data)
- Row 10+: Excel serial-number dates in col 0, float values
- Footnote rows at bottom (dropped)
- Sheet name varies (Data1, Table 1, etc.)

RBA CSV format quirks handled here:
- UTF-8 BOM on first byte
- Row 0: table title
- Row 1: "Title" + column names (actual header)
- Rows 2-10: metadata
- Row 11+: DD-Mon-YYYY dates + values
"""

import re
from pathlib import Path

import numpy as np
import pandas as pd


# ---------------------------------------------------------------------------
# ABS Excel helpers
# ---------------------------------------------------------------------------

def _excel_serial_to_date(serial: float) -> pd.Timestamp:
    """Convert Excel date serial to Timestamp. Excel epoch = 1899-12-30."""
    return pd.Timestamp("1899-12-30") + pd.Timedelta(days=int(serial))


def _clean_abs_column_name(raw: str) -> str:
    """
    Strip ABS series description down to the city/measure name.
    'Residential Property Price Index ;  Sydney ;' -> 'Sydney'
    Returns the last non-empty token between semicolons.
    """
    parts = [p.strip() for p in str(raw).split(";") if p.strip()]
    return parts[-1] if parts else str(raw).strip()


def _find_data_sheet(path: Path) -> str:
    """Return the name of the data sheet in an ABS workbook."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for name in wb.sheetnames:
        if re.match(r"(Data\d+|Table\s*\d+)", name, re.IGNORECASE):
            wb.close()
            return name
    # Fallback to second sheet (index 1) which is usually the data sheet
    result = wb.sheetnames[1] if len(wb.sheetnames) > 1 else wb.sheetnames[0]
    wb.close()
    return result


def parse_abs_excel(
    path: Path,
    n_header_rows: int = 10,
) -> pd.DataFrame:
    """
    Parse an ABS time-series Excel file into a tidy DataFrame.

    Returns DataFrame indexed by pandas Period (quarterly), with one column
    per series (city name extracted from the semicolon-padded header).
    Footnote rows and non-numeric rows are dropped.
    """
    sheet = _find_data_sheet(path)

    # Read raw without header so we control everything
    raw = pd.read_excel(path, sheet_name=sheet, header=None, dtype=object)

    # Extract column names from row 0 (series descriptions)
    col_names = [_clean_abs_column_name(c) for c in raw.iloc[0, 1:].tolist()]

    # Deduplicate: "Sydney", "Sydney" → "Sydney", "Sydney_1"
    seen: dict[str, int] = {}
    unique_col_names = []
    for name in col_names:
        if name in seen:
            seen[name] += 1
            unique_col_names.append(f"{name}_{seen[name]}")
        else:
            seen[name] = 0
            unique_col_names.append(name)
    col_names = unique_col_names

    import datetime as _dt

    # Data starts at row n_header_rows; col 0 = date (datetime object or Excel serial)
    data = raw.iloc[n_header_rows:].reset_index(drop=True)

    # Keep rows where col 0 is a date: either a datetime object or a numeric serial
    def _is_date_cell(val) -> bool:
        if isinstance(val, (_dt.datetime, _dt.date)):
            return True
        try:
            float(val)
            return True
        except (TypeError, ValueError):
            return False

    data = data[data.iloc[:, 0].apply(_is_date_cell)].copy()

    # Convert to Timestamp — handle both datetime objects and Excel serial numbers
    def _to_timestamp(val) -> pd.Timestamp:
        if isinstance(val, (_dt.datetime, _dt.date)):
            return pd.Timestamp(val)
        return _excel_serial_to_date(float(val))

    dates = data.iloc[:, 0].apply(_to_timestamp)
    period_index = pd.PeriodIndex(dates.values, freq="Q-DEC")

    # Build DataFrame from value columns
    values = data.iloc[:, 1:].reset_index(drop=True)
    values.columns = col_names[: values.shape[1]]
    values = values.apply(pd.to_numeric, errors="coerce")
    values.index = period_index

    # Sub-quarterly data (e.g., monthly Labour Force) produces duplicate quarter
    # periods — aggregate to the last observation within each quarter.
    if values.index.duplicated().any():
        values = values.groupby(values.index).last()

    return values


def parse_rppi(path: Path) -> pd.DataFrame:
    """
    Parse ABS 6416.0 RPPI file. Returns DataFrame of price indexes with city columns.
    Keeps only the first series per city (the price index, not the pct-change variants).
    """
    from src.data.config import CITIES

    df = parse_abs_excel(path)

    # Keep first occurrence of each city (deduplication in parse_abs_excel means
    # later variants are "Sydney_1", "Sydney_2" etc. for pct-change series)
    seen_cities: set[str] = set()
    keep = []
    for col in df.columns:
        city_match = next((city for city in CITIES if city == col), None)
        if city_match and city_match not in seen_cities:
            keep.append(col)
            seen_cities.add(city_match)

    # Also keep first weighted average series if present
    for col in df.columns:
        if ("weighted" in col.lower() or "eight" in col.lower()) and col not in keep:
            keep.append(col)
            break

    return df[keep] if keep else df


def parse_cpi(path: Path) -> pd.Series:
    """
    Parse ABS 6401.0 CPI file. Returns the All Groups national CPI series.
    """
    df = parse_abs_excel(path)
    # Look for "All groups" or the first numeric column as national CPI
    for col in df.columns:
        if "all groups" in col.lower() or "all " in col.lower():
            return df[col].dropna().rename("cpi")
    # Fallback: first column
    return df.iloc[:, 0].dropna().rename("cpi")


def parse_labour_force(path: Path) -> pd.Series:
    """
    Parse ABS 6202.0 Labour Force file. Returns seasonally-adjusted unemployment rate.
    The sheet typically has a column containing 'Unemployment rate' with 'Seasonally adjusted'.
    """
    sheet = _find_data_sheet(path)
    raw = pd.read_excel(path, sheet_name=sheet, header=None, dtype=object)

    # Row 0 holds raw series descriptions — search these before they get stripped
    raw_headers = [str(h).lower() for h in raw.iloc[0, 1:].tolist()]

    # Prefer seasonally-adjusted unemployment rate
    target_col_idx = None
    for i, h in enumerate(raw_headers):
        if "unemployment" in h and "rate" in h and "seasonally" in h:
            target_col_idx = i
            break
    if target_col_idx is None:
        for i, h in enumerate(raw_headers):
            if "unemployment" in h and "rate" in h:
                target_col_idx = i
                break

    df = parse_abs_excel(path)

    if target_col_idx is not None:
        # Use iloc to avoid duplicate-column-name issues (df[col] returns DataFrame when names clash)
        return pd.to_numeric(df.iloc[:, target_col_idx], errors="coerce").dropna().rename("unemployment_rate")

    # Last fallback: column whose values sit in the typical unemployment range (2–15%)
    for i in range(df.shape[1]):
        series = pd.to_numeric(df.iloc[:, i], errors="coerce").dropna()
        if len(series) > 20 and series.between(2, 15).mean() > 0.8:
            return series.rename("unemployment_rate")

    raise ValueError(f"Could not find unemployment rate series in {path}")


# ---------------------------------------------------------------------------
# RBA CSV parser
# ---------------------------------------------------------------------------

def parse_rba_csv(path: Path) -> pd.Series:
    """
    Parse RBA F1 cash rate CSV. Returns daily Cash Rate Target series.

    RBA format:
    - UTF-8 BOM on byte 0 (handled by encoding='utf-8-sig')
    - Row 0: table title
    - Row 1: column headers
    - Rows 2-10: metadata
    - Row 11+: DD-Mon-YYYY dates + values
    """
    # Read with BOM stripping; row 1 is the real header
    raw = pd.read_csv(
        path,
        encoding="utf-8-sig",
        header=1,
        skiprows=[2, 3, 4, 5, 6, 7, 8, 9, 10],
        low_memory=False,
    )

    # First column is the date column
    date_col = raw.columns[0]
    raw = raw.dropna(subset=[date_col])
    raw = raw[raw[date_col].astype(str).str.match(r"\d{2}-\w{3}-\d{4}")]

    dates = pd.to_datetime(raw[date_col], format="%d-%b-%Y")

    # Find Cash Rate Target column
    target_col = None
    for col in raw.columns[1:]:
        if "cash rate target" in col.lower():
            target_col = col
            break
    if target_col is None:
        # Try first numeric column as fallback
        for col in raw.columns[1:]:
            vals = pd.to_numeric(raw[col], errors="coerce")
            if vals.notna().sum() > 50:
                target_col = col
                break

    if target_col is None:
        raise ValueError(f"Could not find cash rate column in {path}")

    values = pd.to_numeric(raw[target_col], errors="coerce")
    series = pd.Series(values.values, index=dates, name="cash_rate")
    return series.dropna().sort_index()


# ---------------------------------------------------------------------------
# Resampling and merging
# ---------------------------------------------------------------------------

def resample_to_quarter(series: pd.Series, method: str = "last") -> pd.Series:
    """
    Resample a daily/monthly series to quarterly frequency (Q-DEC).
    Uses last-in-quarter value by default (end-of-quarter rate = known rate for that period).
    Returns PeriodIndex series.
    """
    freq = "Q-DEC" if isinstance(series.index, pd.PeriodIndex) else "QE-DEC"
    resampled = series.resample(freq).last() if method == "last" else series.resample(freq).mean()
    period_idx = pd.PeriodIndex(resampled.index, freq="Q-DEC")
    result = pd.Series(resampled.values, index=period_idx, name=series.name)
    return result.dropna()


def merge_to_panel(
    rppi: pd.DataFrame,
    cash_rate: pd.Series,
    cpi: pd.Series,
    unemployment: pd.Series,
) -> pd.DataFrame:
    """
    Build a long-format panel DataFrame: (city, period) rows.

    All series are aligned on quarterly PeriodIndex. Inner join keeps only
    quarters where all sources have data.

    Returns DataFrame with columns:
        city, period, rppi_index, cash_rate, cpi, unemployment_rate
    """
    # Ensure all macro series are quarterly
    if not isinstance(cash_rate.index, pd.PeriodIndex):
        cash_rate = resample_to_quarter(cash_rate)
    if not isinstance(cpi.index, pd.PeriodIndex):
        cpi = resample_to_quarter(cpi)
    if not isinstance(unemployment.index, pd.PeriodIndex):
        unemployment = resample_to_quarter(unemployment)

    # Align macros on common quarters
    macro = pd.DataFrame(
        {"cash_rate": cash_rate, "cpi": cpi, "unemployment_rate": unemployment}
    ).dropna()

    # Melt RPPI from wide to long
    from src.data.config import CITIES

    city_cols = [c for c in rppi.columns if any(city in c for city in CITIES)]
    rows = []
    for col in city_cols:
        city_name = next((city for city in CITIES if city in col), col)
        city_series = rppi[col].dropna()
        merged = city_series.rename("rppi_index").to_frame().join(macro, how="inner")
        merged["city"] = city_name
        merged.index.name = "period"
        merged = merged.reset_index()
        rows.append(merged)

    panel = pd.concat(rows, ignore_index=True)
    panel = panel.sort_values(["city", "period"]).reset_index(drop=True)
    return panel
